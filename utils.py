from curl_cffi import requests
from settings import headers
from bs4 import BeautifulSoup
import re
import json
import openpyxl


class PropertyOfProduct:
    def __init__(self):
        self.lenght = None
        self.width = None
        self.form = None
        self.material = None
        self.hydromassage = None
        self.spilover = None
        self.hendles = None
        self.country = None
        self.deep = None
        self.height_support = None
        self.montage = None
        self.guarantee = None
        self.volume = None
        self.style = None
        self.cover = None

    def get_property(self, data):
        for params in data:
            for param in params["values"]:
                if "Длина" in param["prop"]: self.lenght = param["value"]
                if "Ширина" in param["prop"]: self.width = param["value"]
                if "Форма" in param["prop"]: self.form = param["value"]
                if "Материал" in param["prop"]: self.material = param["value"]
                if "Система гидромассажа" in param["prop"]: self.hydromassage = param["value"]
                if "Расположение перелива" in param["prop"]: self.spilover = param["value"]
                if "Ручки" in param["prop"]: self.hendles = param["value"]
                if "Страна" in param["prop"]: self.country = param["value"]
                if "Глубина" in param["prop"]: self.deep = param["value"]
                if "Высота с опорой" in param["prop"]: self.height_support = param["value"]
                if "Монтаж" in param["prop"]: self.montage = param["value"]
                if "Гарантия" in param["prop"]: self.guarantee = param["value"]
                if "Объем" in param["prop"]: self.volume = param["value"]
                if "Стиль" in param["prop"]: self.style = param["value"]
                if "Поверхность" in param["prop"]: self.cover = param["value"]

    def get_all_property(self):
        all_propertys = [self.lenght, self.width, self.form, self.material, self.hydromassage,
                         self.spilover, self.hendles, self.country, self.deep, self.height_support, self.montage,
                         self.guarantee, self.volume, self.style, self.cover]
        return all_propertys

    @staticmethod
    def get_all_links(category_link):
        page = 1
        all_links = []
        while True:
            data = {
                'PAGEN_1': page,
                'perpage': '71',
                'sort_field': 'SORT',
                'sort_order': 'DESC',
                'sectionId': '7697',
            }
            response = requests.get(category_link, impersonate="chrome99", max_redirects=3, headers=headers,
                                    data=data).json()
            data = response.get("data")
            links_perpage = []
            for product in data:
                link = product["link"]
                title = product["title"]
                brand = product.get("brand")
                add_link = f"https://santehnika-online.ru{link}"
                links_perpage.append([add_link, title, brand])
            all_links.extend(links_perpage)
            print(f"Получена страница № {page}")
            if len(data) != 71:
                break
            page += 1
        return all_links

    @staticmethod
    def get_product(link):
        while True:
            response = requests.get(link, impersonate="chrome110", max_redirects=3, headers=headers)
            soup = BeautifulSoup(response.text, "lxml")
            pattern = r'var __SD__ = {"Location":{"option":.*?};'
            matches = re.findall(pattern, str(soup))
            for match in matches:
                json_match = re.search(r'\{.*\}', match).group()
                data_dict = json.loads(json_match)
            try:
                data = data_dict.get("CardProductSpec").get("data")[0].get("groups").get("additional")
                return data
            except UnboundLocalError:
                print("Ошибка. Отправляю запрос ещще раз.")

    @staticmethod
    def first_sheet(sheet):
        sheet.cell(row=1, column=1, value="Назввание товара")
        sheet.cell(row=1, column=2, value="Ссылка на товар")
        sheet.cell(row=1, column=3, value="Брэнд")
        sheet.cell(row=1, column=4, value="Длина")
        sheet.cell(row=1, column=5, value="Ширина")
        sheet.cell(row=1, column=6, value="Форма")
        sheet.cell(row=1, column=7, value="Материал")
        sheet.cell(row=1, column=8, value="Система гидромассажа")
        sheet.cell(row=1, column=9, value="Расположение перелива")
        sheet.cell(row=1, column=10, value="Ручки")
        sheet.cell(row=1, column=11, value="Страна")
        sheet.cell(row=1, column=12, value="Глубина")
        sheet.cell(row=1, column=13, value="Высота с опорой")
        sheet.cell(row=1, column=14, value="Монтаж")
        sheet.cell(row=1, column=15, value="Гарантия")
        sheet.cell(row=1, column=16, value="Объем")
        sheet.cell(row=1, column=17, value="Стиль")
        sheet.cell(row=1, column=18, value="Поверхность")

    @staticmethod
    def save_to_excel(title, link, brand, all_propertys, row, sheet):
        sheet.cell(row=row, column=1, value=title)
        sheet.cell(row=row, column=2, value=link)
        sheet.cell(row=row, column=3, value=brand)
        for i, product_property in enumerate(all_propertys, start=4):
            sheet.cell(row=row, column=i, value=product_property)

    @staticmethod
    def get_parsing_brands(path):
        brands = []
        with open(path, "r", encoding="utf-8") as file:
            lines = file.readlines()
        for line in lines:
            brand = "".join(line.split("—")[0]).strip()
            brands.append(brand.lower())
        return brands
